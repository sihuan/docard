import redis, requests, json
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill,colors,Fill

from config import APIURL, AUTHORIZATION

pool = redis.ConnectionPool(host='127.0.0.1', port=6379, decode_responses=True)
r = redis.Redis(connection_pool=pool)

class studentProfile():

    def __init__(self,sid,name,major,classroom):
        self.sid = str(sid)
        self.name = name
        self.major = major
        self.classroom = classroom
        self.normal = '无异常'
        self.nowArea = ''
        self.te = '正常'
        self.knowNoReturn = '知晓'
        self.changeArea = '无'
        self.video = '知道'

def addNewStudent(student):
    r.sadd('allStudentSet',student.sid)
    r.hmset('student' + student.sid, student.__dict__)

def doCard(sid,nowArea,te,knowNoReturn,changeArea,video):
    sid = str(sid)
    a = list(r.smembers("allStudentSet"))
    if sid in a:
        if te != '正常' or knowNoReturn != '知晓' or changeArea != '否' or video != '知道':
            normal = '异常'
        else:
            normal = '无异常'
        r.hmset('student' + sid,{
            'normal':normal,
            'nowArea':nowArea,
            'te':te,
            'knowNoReturn':knowNoReturn,
            'changeArea':changeArea,
            'video': video
        })
        r.sadd('doCardStudent',sid)
        return True
    else:
        return False


def loadStudent(filename):
    wb = load_workbook(filename=filename)
    ws = wb.active
    rows = ws.rows
    for row in rows:
        line = [col.value for col in row]
        newstudent = studentProfile(line[0],line[1],line[2],line[3])
        addNewStudent(newstudent)

def export(filename):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "已填表"
    ws2 = wb.create_sheet(title='未填表')
    ws3 = wb.create_sheet(title='异常')

    redfill = PatternFill('solid', fgColor='CB1B45')
    yellowfill = PatternFill('solid', fgColor='FFB11B')
    okstatus = {
        7:'正常',
        8:'知晓',
        9:'否',
        10:'知道',
    }
    ws1.column_dimensions['A'].width = 20.0
    ws1.column_dimensions['D'].width = 20.0
    ws1.column_dimensions['E'].width = 20.0
    ws1.column_dimensions['F'].width = 20.0
    ws1.column_dimensions['H'].width = 20.0
    ws1.column_dimensions['I'].width = 20.0

    ws3.column_dimensions['A'].width = 20.0
    ws3.column_dimensions['D'].width = 20.0
    ws3.column_dimensions['E'].width = 20.0
    ws3.column_dimensions['F'].width = 20.0
    ws3.column_dimensions['H'].width = 20.0
    ws3.column_dimensions['I'].width = 20.0

    ws2.column_dimensions['A'].width = 20.0
    ws2.column_dimensions['D'].width = 20.0


    ws1.append(['学号','姓名','专业','班级','是否存在异常','居住地','体温','是否知晓不可随意反青','是否有地址迁移','是否知道疫情防控节目'])
    ws2.append(['学号','姓名','专业','班级'])
    ws3.append(['学号','姓名','专业','班级','是否存在异常','居住地','体温','是否知晓不可随意反青','是否有地址迁移','是否知道疫情防控节目'])
    
    allstudent = list(r.smembers("allStudentSet"))
    allstudent.sort()
    docardstudent = list(r.smembers("doCardStudent"))
    docardstudent.sort()

    for sid in allstudent:
        ts = list(r.hgetall('student' + sid).values())

        if sid in docardstudent:
            if ts[4] == '无异常':
                ws1.append(ts)
            else:
                ws1.append(ts)
                ws3.append(ts)
                r1 = ws1.max_row
                r3 = ws3.max_row

                ws1.cell(r1,5).fill = redfill
                ws3.cell(r3,5).fill = redfill
                for c in range(7,11):
                    if ts[c-1] != okstatus[c]:
                        ws1.cell(r1,c).fill = yellowfill
                        ws3.cell(r3,c).fill = yellowfill

        else:
            ws2.append(ts[:4])
        
    wb.save(filename)

def findstudent(sid):
    sid = 'student'+str(sid)
    return r.hgetall(sid)

def checkdata(cr):
    allstudent = list(r.smembers("allStudentSet"))
    allstudent.sort()
    docardstudent = list(r.smembers("doCardStudent"))
    docardstudent.sort()

    yidaka=[]
    weidaka=[]
    yichang=[]

    for sid in allstudent:
        ts = r.hmget('student' + sid,'classroom','name','normal')
        if cr == ts[0]:
            if sid in docardstudent:
                if ts[2] == '无异常':
                    yidaka.append(ts[1])
                else:
                    yidaka.append(ts[1])
                    yichang.append(ts[1])
            else:
                weidaka.append(ts[1])
    return {
        'ydk':yidaka,
        'wdk':weidaka,
        'yc':yichang,
    }
            
def tipall():
    crs = {
        '采矿工程2018-1班':[[1687402949,467701053,],'采矿工程2018-1班未打卡的有：'],
        '采矿工程2018-2班':[[1565638695],'采矿工程2018-2班未打卡的有：'],
        '采矿工程2018-3班':[[2633876343,],'采矿工程2018-3班未打卡的有：'],
        '采矿工程2018-4班':[[3218982708,1134911219],'采矿工程2018-4班未打卡的有：'],
        '采矿工程2018-5班':[[3140974899,1421394778,],'采矿工程2018-5班未打卡的有：'],
        '采矿工程2018-6班':[[1225353152,1787559758,1079518153],'采矿工程2018-6班未打卡的有：'],
        '工业工程2018-1班':[[1244486741,],'工业工程2018-1班未打卡的有：'],
        '工业工程2018-3班':[[1834650704,],'工业工程2018-3班未打卡的有：'],
        '工业工程2018-2班':[[2638301362,941279995],'工业工程2018-2班未打卡的有：'],
        '工程力学2018-1班':[[772290137,1961569211],'工程力学2018-1班未打卡的有：'],
        '工程力学2018-2班':[[1805616132,1005531288],'工程力学2018-2班未打卡的有：'],
        '工程力学2018-3班':[[1064486615,],'工程力学2018-3班未打卡的有：'],
    }

    # allstudent = list(r.smembers("allStudentSet"))
    # allstudent.sort()
    # docardstudent = list(r.smembers("doCardStudent"))
    # docardstudent.sort()
    
    notdocard = list(r.sdiff("allStudentSet", "doCardStudent"))
    notdocard.sort()

    print(notdocard)

    for sid in notdocard:
        ts = r.hmget('student' + sid,'classroom','name')
        crs[ts[0]][1] = crs[ts[0]][1] + '\n' + ts[1]

    for crn, cr in crs.items():
        print(crn)
        if len(cr[1]) == 17:
            cr[1] = '本班已经全部打卡完毕。' 


        for banzhang in cr[0]:
            status = 'notok'
            for i in range(3): 
                status = sendmsg(cr[1],banzhang)
                print(crn+status)
                if status == 'ok':
                    break
            sendmsg(crn+status+str(banzhang),1318000868)


def sendmsg(msg,uid):
    url = APIURL + "send_msg"
    Headers = {
        'content-type': 'application/json',
        'Authorization':'Bearer ' + AUTHORIZATION
        }
    
    data = {
        'message_type' : 'private',
        'user_id' : uid,
        'message': msg
        }
    try:
        resp = requests.post(url = url, data = json.dumps(data),headers = Headers)
        print(resp.json()['status'])
        return resp.json()['status']
    except Exception as e:
        return str(e)

def checkalldata():
    allstudent = list(r.smembers("allStudentSet"))
    allstudent.sort()
    docardstudent = list(r.smembers("doCardStudent"))
    docardstudent.sort()

    wdk=[]
    yc=[]

    for sid in allstudent:
        ts = r.hmget('student' + sid,'normal','name','classroom','te','knowNoReturn','changeArea','video')
        if sid in docardstudent:
            if ts[0] != '无异常':
                yc.append({
                    'name':ts[1],
                    'cr':ts[2],
                    'te':ts[3],
                    'nr':ts[4],
                    'ch':ts[5],
                    'video':ts[6],
                })
        else:
            wdk.append({
                'name':ts[1],
                'cr':ts[2]
            })

    return {
        'wdk':wdk,
        'yc':yc,
    }